#!/usr/bin/env python3
"""
AUROS AI — Batch Lead Runner
Scrapes 50+ leads/day across all Georgia cities and multiple business types.
Outputs a clean Excel pipeline tracker for cold calling.

Usage:
    python tools/batch_lead_runner.py --target 50
    python tools/batch_lead_runner.py --target 50 --types "salon,gym,contractor"
    python tools/batch_lead_runner.py --target 50 --cities "Savannah,Augusta"
"""

from __future__ import annotations

import sys
import json
import argparse
import logging
import os
import time
from datetime import datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from dotenv import load_dotenv
load_dotenv(PROJECT_ROOT / ".env", override=True)

from tools.lead_scraper import search_businesses

logger = logging.getLogger("auros.batch")

# ─── Georgia Cities (by population, top 30) ───────────────────────────────
GEORGIA_CITIES = [
    "Atlanta", "Augusta", "Columbus", "Macon", "Savannah",
    "Athens", "Sandy Springs", "Roswell", "Johns Creek", "Albany",
    "Warner Robins", "Alpharetta", "Marietta", "Valdosta", "Smyrna",
    "Dunwoody", "Brookhaven", "Peachtree City", "Newnan", "Dalton",
    "Gainesville", "Milton", "Hinesville", "Statesboro", "Lawrenceville",
    "Duluth", "Kennesaw", "Woodstock", "Canton", "Carrollton",
]

# ─── Business Types That Need Websites Most ───────────────────────────────
BUSINESS_TYPES = [
    "restaurant",
    "hair salon",
    "barber shop",
    "nail salon",
    "auto repair",
    "dentist",
    "plumber",
    "electrician",
    "landscaping",
    "cleaning service",
    "gym fitness",
    "yoga studio",
    "bakery",
    "cafe coffee shop",
    "pet grooming",
    "tattoo shop",
    "florist",
    "photographer",
    "daycare",
    "church",
    "contractor",
    "roofing",
    "hvac",
    "accounting",
    "chiropractor",
    "massage spa",
    "dog trainer",
    "tutoring",
    "tailor alterations",
    "car wash",
]


def run_batch(
    target: int = 50,
    cities: list[str] | None = None,
    business_types: list[str] | None = None,
    min_rating: float = 4.0,
    min_reviews: int = 15,
) -> list[dict]:
    """
    Run batch scraping across cities and business types until we hit our target.
    Rotates through city+type combos to get diverse leads.
    """
    use_cities = cities or GEORGIA_CITIES
    use_types = business_types or BUSINESS_TYPES

    all_leads = []
    seen_place_ids = set()
    api_calls = 0
    max_api_calls = 300  # Safety cap to avoid burning through quota

    logger.info(f"Target: {target} leads across {len(use_cities)} cities, {len(use_types)} business types")

    # Create combos and rotate through them
    combos = []
    for btype in use_types:
        for city in use_cities:
            combos.append((btype, city))

    for btype, city in combos:
        if len(all_leads) >= target:
            break
        if api_calls >= max_api_calls:
            logger.warning(f"Hit API call safety cap ({max_api_calls}). Stopping.")
            break

        query = f"{btype} in {city} GA"
        logger.info(f"Searching: {query} ({len(all_leads)}/{target} leads so far)")

        try:
            # Only grab a few per combo to diversify
            results = search_businesses(
                query=query,
                max_results=10,
                min_rating=min_rating,
                min_reviews=min_reviews,
                filter_no_website=True,
            )
            api_calls += 1 + len(results)  # 1 for search + 1 per detail lookup

            for biz in results:
                pid = biz.get("place_id", "")
                if pid and pid not in seen_place_ids:
                    seen_place_ids.add(pid)
                    biz["search_category"] = btype
                    biz["search_city"] = city
                    biz["scraped_date"] = datetime.now().strftime("%Y-%m-%d")
                    # Add pipeline fields
                    biz["pipeline_status"] = "New"
                    biz["call_status"] = "Not Called"
                    biz["call_notes"] = ""
                    biz["interest_level"] = ""
                    biz["meeting_date"] = ""
                    biz["website_built"] = "No"
                    biz["deal_value"] = ""
                    biz["deal_status"] = ""
                    all_leads.append(biz)

                    if len(all_leads) >= target:
                        break

            # Small delay between searches to be respectful
            time.sleep(0.5)

        except Exception as e:
            logger.error(f"Error searching {query}: {e}")
            continue

    logger.info(f"Batch complete: {len(all_leads)} leads found, {api_calls} API calls used")
    return all_leads


def export_to_excel(leads: list[dict], filename: str = "") -> str:
    """Export leads to a clean Excel pipeline tracker."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        return ""

    if not filename:
        date_str = datetime.now().strftime("%Y-%m-%d")
        filename = f"AUROS_Lead_Pipeline_{date_str}.xlsx"

    filepath = PROJECT_ROOT / "portfolio" / filename
    filepath.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ── Sheet 1: Active Pipeline ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Pipeline"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Status color fills
    status_colors = {
        "Not Called": PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"),
        "Called - No Answer": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "Called - Not Interested": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        "Called - Interested": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        "Meeting Booked": PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
        "Website Built": PatternFill(start_color="D1C4E9", end_color="D1C4E9", fill_type="solid"),
        "SOLD": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
        "Lost": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
    }

    row_alt_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )

    # Headers
    headers = [
        "#",
        "Business Name",
        "Category",
        "City",
        "Phone",
        "Rating",
        "Reviews",
        "Google Maps",
        "Pipeline Status",
        "Call Status",
        "Call Notes",
        "Interest (1-5)",
        "Meeting Date",
        "Website Built?",
        "Deal Value ($)",
        "Deal Status",
        "Date Added",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:Q{len(leads) + 1}"

    # Data rows
    for i, lead in enumerate(leads, 1):
        row = i + 1
        data = [
            i,
            lead.get("business_name", ""),
            lead.get("search_category", lead.get("category", "")),
            lead.get("city", lead.get("search_city", "")),
            lead.get("phone", ""),
            lead.get("rating", ""),
            lead.get("review_count", ""),
            lead.get("google_maps_url", ""),
            lead.get("pipeline_status", "New"),
            lead.get("call_status", "Not Called"),
            lead.get("call_notes", ""),
            lead.get("interest_level", ""),
            lead.get("meeting_date", ""),
            lead.get("website_built", "No"),
            lead.get("deal_value", ""),
            lead.get("deal_status", ""),
            lead.get("scraped_date", datetime.now().strftime("%Y-%m-%d")),
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col in [11]))

            # Alternating row colors
            if i % 2 == 0:
                cell.fill = row_alt_fill

            # Bold the business name
            if col == 2:
                cell.font = Font(bold=True, size=11)

            # Color the rating
            if col == 6 and isinstance(value, (int, float)):
                if value >= 4.8:
                    cell.font = Font(color="1B5E20", bold=True)
                elif value >= 4.5:
                    cell.font = Font(color="33691E")

            # Make Google Maps a hyperlink
            if col == 8 and value:
                cell.hyperlink = value
                cell.font = Font(color="0066CC", underline="single")
                cell.value = "Open Map"

    # Column widths
    col_widths = [5, 30, 18, 16, 16, 8, 9, 12, 18, 22, 30, 12, 14, 14, 14, 14, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Sheet 2: Pipeline Summary ─────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")

    summary_header_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")

    ws2.cell(row=1, column=1, value="AUROS AI — Lead Pipeline Summary").font = Font(bold=True, size=16, color="1A1A2E")
    ws2.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}").font = Font(size=10, color="666666")

    # Stats
    stats = [
        ("Total Leads", len(leads)),
        ("Cities Covered", len(set(l.get("city", l.get("search_city", "")) for l in leads))),
        ("Business Types", len(set(l.get("search_category", "") for l in leads))),
        ("Avg Rating", round(sum(l.get("rating", 0) for l in leads) / max(len(leads), 1), 1)),
        ("Avg Reviews", round(sum(l.get("review_count", 0) for l in leads) / max(len(leads), 1))),
    ]

    ws2.cell(row=4, column=1, value="Metric").font = header_font
    ws2.cell(row=4, column=1).fill = summary_header_fill
    ws2.cell(row=4, column=2, value="Value").font = header_font
    ws2.cell(row=4, column=2).fill = summary_header_fill

    for i, (metric, value) in enumerate(stats, 5):
        ws2.cell(row=i, column=1, value=metric).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=value)

    # Breakdown by city
    ws2.cell(row=11, column=1, value="Leads by City").font = Font(bold=True, size=12)
    city_counts = {}
    for lead in leads:
        city = lead.get("city", lead.get("search_city", "Unknown"))
        city_counts[city] = city_counts.get(city, 0) + 1

    ws2.cell(row=12, column=1, value="City").font = header_font
    ws2.cell(row=12, column=1).fill = summary_header_fill
    ws2.cell(row=12, column=2, value="Count").font = header_font
    ws2.cell(row=12, column=2).fill = summary_header_fill

    for i, (city, count) in enumerate(sorted(city_counts.items(), key=lambda x: -x[1]), 13):
        ws2.cell(row=i, column=1, value=city)
        ws2.cell(row=i, column=2, value=count)

    # Breakdown by type
    type_start = 13 + len(city_counts) + 2
    ws2.cell(row=type_start, column=1, value="Leads by Business Type").font = Font(bold=True, size=12)
    type_counts = {}
    for lead in leads:
        btype = lead.get("search_category", "Unknown")
        type_counts[btype] = type_counts.get(btype, 0) + 1

    ws2.cell(row=type_start + 1, column=1, value="Type").font = header_font
    ws2.cell(row=type_start + 1, column=1).fill = summary_header_fill
    ws2.cell(row=type_start + 1, column=2, value="Count").font = header_font
    ws2.cell(row=type_start + 1, column=2).fill = summary_header_fill

    for i, (btype, count) in enumerate(sorted(type_counts.items(), key=lambda x: -x[1]), type_start + 2):
        ws2.cell(row=i, column=1, value=btype)
        ws2.cell(row=i, column=2, value=count)

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15

    # ── Sheet 3: Call Script Reference ────────────────────────────────────
    ws3 = wb.create_sheet("Call Script")

    ws3.cell(row=1, column=1, value="30-SECOND COLD CALL SCRIPT").font = Font(bold=True, size=16, color="1A1A2E")
    ws3.column_dimensions["A"].width = 100

    script_lines = [
        "",
        "━━━ WHEN THEY PICK UP ━━━",
        "",
        '"Hi, is this [OWNER NAME / the owner]?"',
        "",
        '"Hey [NAME], my name is Leo — I run a small web design company."',
        '"I was looking at businesses in [CITY] on Google and I noticed"',
        '"you guys have amazing reviews — [RATING] stars, [REVIEWS] reviews —"',
        '"but no website. I was curious, have you ever thought about getting one?"',
        "",
        "━━━ IF YES / MAYBE ━━━",
        "",
        '"Cool — I actually specialize in building sites for local businesses"',
        '"like yours. What I\'d love to do is set up a quick 15-minute call"',
        '"where I can show you what a site for [BUSINESS] could look like."',
        '"I\'ll even put something together for free so you can see it"',
        '"before you commit to anything. When works better for you,"',
        '"this week or early next week?"',
        "",
        "━━━ IF NO / NOT INTERESTED ━━━",
        "",
        '"No worries at all. Just thought I\'d reach out since your reviews"',
        '"are really strong and a website would help more people find you."',
        '"If you ever change your mind, I\'m Leo at AUROS — have a great day!"',
        "",
        "━━━ IF VOICEMAIL ━━━",
        "",
        '"Hey [NAME], this is Leo. I run a web design company and I noticed"',
        '"[BUSINESS] has incredible Google reviews but no website."',
        '"I\'d love to chat about putting one together for you."',
        '"My number is [YOUR NUMBER]. Thanks!"',
        "",
        "━━━ OBJECTION HANDLERS ━━━",
        "",
        '"We get all our business from word of mouth"',
        '→ "That\'s awesome — imagine how much more you\'d get if people"',
        '  "could actually find you online too. 97% of people search"',
        '  "Google before visiting a local business."',
        "",
        '"We can\'t afford a website"',
        '→ "Totally understand. I\'m actually offering to build the first"',
        '  "version for free — no cost, no commitment. If you love it,"',
        '  "we can talk about a small monthly fee to keep it running."',
        "",
        '"We\'re too busy right now"',
        '→ "That\'s actually the best time — a website works for you 24/7."',
        '  "It takes zero effort on your end. Can I send you a quick"',
        '  "preview and you check it when you have a minute?"',
        "",
        "━━━ KEY STATS TO DROP ━━━",
        "",
        "• 97% of consumers search online before visiting a local business",
        "• 75% judge a business's credibility based on their website",
        "• Businesses with websites get 2.5x more customers on average",
        "• Google prioritizes businesses WITH websites in local search results",
    ]

    for i, line in enumerate(script_lines, 2):
        cell = ws3.cell(row=i, column=1, value=line)
        if line.startswith("━━━"):
            cell.font = Font(bold=True, size=12, color="C08B5C")
        elif line.startswith('"') and "→" not in line:
            cell.font = Font(size=11, italic=True)
        elif line.startswith("→"):
            cell.font = Font(size=11, color="1B5E20")
        elif line.startswith("•"):
            cell.font = Font(size=11, bold=True)

    # Save
    wb.save(str(filepath))
    logger.info(f"Excel saved: {filepath}")
    return str(filepath)


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(name)s] %(message)s")

    parser = argparse.ArgumentParser(description="AUROS Batch Lead Runner")
    parser.add_argument("--target", type=int, default=50, help="Target number of leads")
    parser.add_argument("--cities", type=str, help="Comma-separated cities (default: all GA cities)")
    parser.add_argument("--types", type=str, help="Comma-separated business types (default: all)")
    parser.add_argument("--min-rating", type=float, default=4.0)
    parser.add_argument("--min-reviews", type=int, default=15)
    parser.add_argument("--output", type=str, default="", help="Output filename")

    args = parser.parse_args()

    cities = [c.strip() for c in args.cities.split(",")] if args.cities else None
    types = [t.strip() for t in args.types.split(",")] if args.types else None

    print(f"\n🔍 AUROS AI Lead Runner")
    print(f"   Target: {args.target} leads")
    print(f"   Cities: {', '.join(cities) if cities else 'All 30 Georgia cities'}")
    print(f"   Types: {', '.join(types) if types else 'All 30 business types'}")
    print(f"   Min Rating: {args.min_rating}★ | Min Reviews: {args.min_reviews}")
    print()

    leads = run_batch(
        target=args.target,
        cities=cities,
        business_types=types,
        min_rating=args.min_rating,
        min_reviews=args.min_reviews,
    )

    if leads:
        filepath = export_to_excel(leads, filename=args.output)
        print(f"\n✅ {len(leads)} leads exported to: {filepath}")
        print(f"   Open it up and start calling!")
    else:
        print("\n❌ No leads found. Check your API key or try different search params.")
